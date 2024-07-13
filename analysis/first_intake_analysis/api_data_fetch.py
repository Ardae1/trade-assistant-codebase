import requests
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import datetime
from sklearn import linear_model
import concurrent.futures
import scipy.stats as stats
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from configs.config import Config
from functools import partial


config = Config()


class BinanceDataHelper:
    def __init__(self):
        self.header = {"X-MBX-APIKEY": f"{config.BINANCE_API_KEY_PATH}"}

    def convert_time(self, time):
        dt = pd.to_datetime(time)
        dt = int(dt.timestamp())
        dt = dt * 1000
        print("conver time başarılı..")
        return dt

    def reconv_date(self, time):
        time = time / 1000
        dt_obj = datetime.datetime.fromtimestamp(time)
        print("recconver time başarılı..")
        return dt_obj.strftime("%Y-%m-%d %H:%M:%S")

    def calc_values(self, df):
        pplusrange = []
        pnegrange = []
        nplusrange = []
        nnegrange = []

        for index, group in df.iterrows():
            if (
                isinstance(group["close"], (int, float))
                and isinstance(group["Open"], (int, float))
                and isinstance(group["High"], (int, float))
                and isinstance(group["low"], (int, float))
            ):
                if group["close"] > group["Open"]:
                    pvalue = group["High"] - group["Open"]
                    nvalue = group["Open"] - group["low"]
                    pvalue = (pvalue / group["Open"]) * 100
                    nvalue = (nvalue / group["Open"]) * 100
                    pplusrange.append(pvalue)
                    pnegrange.append(nvalue)
                else:
                    pplusrange.append(0)
                    pnegrange.append(0)

                if group["close"] < group["Open"]:
                    pvalue = group["High"] - group["Open"]
                    nvalue = group["Open"] - group["low"]
                    pvalue = (pvalue / group["Open"]) * 100
                    nvalue = (nvalue / group["Open"]) * 100
                    nplusrange.append(pvalue)
                    nnegrange.append(nvalue)
                else:
                    nplusrange.append(0)
                    nnegrange.append(0)
            else:
                print(f"Invalid data types at index {index}")
                pplusrange.append(0)
                pnegrange.append(0)
                nplusrange.append(0)
                nnegrange.append(0)

        print("calc values başarılı..")
        return pplusrange, pnegrange, nplusrange, nnegrange

    def add_values_dframe(self, df, c1, c2, c3, c4):
        df["pplusrange"] = c1
        df["pnegrange"] = c2
        df["nplusrange"] = c3
        df["nnegrange"] = c4
        print("add_values başarılı..")
        return df

    def adj_dataframe(self, df):
        columns = ["time", "Open", "High", "low", "close"]
        df = df.drop(df.columns[5:12], axis=1)
        df = df.rename(columns={0: "time", 1: "Open", 2: "High", 3: "low", 4: "close"})
        df["time"] = df["time"].apply(self.reconv_date)
        df.set_index("time", inplace=True)
        df.index = pd.to_datetime(df.index)

        # Convert columns to float
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        df["dayrange"] = df["High"] - df["low"]
        df["diff"] = df["dayrange"] / df["close"]
        df["diff"] = df["diff"] * 100

        c1, c2, c3, c4 = self.calc_values(df)
        df = self.add_values_dframe(df, c1, c2, c3, c4)
        print("adj dataframe başarılı..")
        return df

    def params_toJson(
        self, parity, contract_type="spot", tf="15m", unix_time=0, count_index=0
    ):
        start_time = count_index
        end_time = count_index + unix_time
        if contract_type == "spot":
            params = {
                "symbol": parity,
                "interval": tf,
                "startTime": start_time,
                "endTime": end_time,
            }
        else:
            params = {
                "symbol": parity,
                "contractType": contract_type,
                "interval": tf,
                "startTime": start_time,
                "endTime": end_time,
            }
        print("json başarılı..")
        return params

    def send_request(self, parity, type, tf, unix_time, count_index):
        params = self.params_toJson(parity, type, tf, unix_time, count_index)
        if type == "spot":
            url = config.BINANCE_API_SPOT_URL
        else:
            url = config.BINANCE_API_FUTURE_URL

        try:
            data = requests.get(url, headers=self.header, params=params)
            data.raise_for_status()  # Raise HTTPError for bad responses
            data = data.json()
            print("main request başarılı..")
            data = pd.DataFrame(data)
            data = self.adj_dataframe(data)
            data.drop(data.index[-1], inplace=True)
            print("api request başarılı..")
            return data
        except requests.exceptions.HTTPError as http_err:
            print(f"HTTP error occurred: {http_err}")
            raise
        except Exception as e:
            print(f"Error processing data for timeframe {tf}: {e}")
            raise

    def unix_time_indexer(self, tf: str):
        if tf == "15m":
            unix_indexer = 432000000
        elif tf == "1h":
            unix_indexer = 1728000000
        elif tf == "4h":
            unix_indexer = 7156800000
        else:
            unix_indexer = 0
        print("unix_time başarılı..")
        return unix_indexer

    def time_setter(self, start, end=None, index=None):
        start_date = self.convert_time(start)
        if end is None:
            today = self.convert_time(datetime.datetime.now().strftime("%Y-%m-%d"))
        else:
            today = self.convert_time(end)
        print("time setter başarılı..")
        return np.arange(start_date, today, index)

    def main(self, coin, contract_type="spot", tf="15m", start=None, end=None):
        main_list = []

        dbliste = []
        unix_index = self.unix_time_indexer(tf)
        time_range = self.time_setter(start, end, unix_index)
        for time_index in time_range:
            requested_data = self.send_request(
                coin, contract_type, tf, unix_index, time_index
            )
            dbliste.append(requested_data)

        dblast = pd.concat(dbliste, axis=0)
        dblast.reset_index(inplace=True)
        dblast.insert(0, "parity", coin)
        dblast.set_index("parity", inplace=True)
        print("Coin:", coin, "prepared..")
        main_list.append(dblast)

        return pd.concat(main_list)

    def excel_mapper(self, results):
        for result in results:
            for sub_result in result:
                check_date_1 = pd.to_datetime(sub_result.iloc[0]["time"])
                check_date_2 = pd.to_datetime(sub_result.iloc[1]["time"])
                example_name = sub_result.index[0]

                date_condition = (check_date_2 - check_date_1).total_seconds() / 60

                self.excel_time_adapter(str(date_condition), example_name, sub_result)

        return 1

    def excel_time_adapter(self, minute, name, result):
        time_dict = {
            "15.0_usdt": "/Users/ardaerkan/Documents/MİGRATE/Python/Trade_Assistant_Files/crypto-15min-dataset.csv",
            "15.0_btc": "/Users/ardaerkan/Documents/MİGRATE/Python/Trade_Assistant_Files/crypto_15min_dataset_BTC.csv",
            "60.0_usdt": "/Users/ardaerkan/Documents/MİGRATE/Python/Trade_Assistant_Files/crypto-1h-dataset.csv",
            "60.0_btc": "/Users/ardaerkan/Documents/MİGRATE/Python/Trade_Assistant_Files/crypto_1h_dataset_BTC.csv",
            "240.0_usdt": "/Users/ardaerkan/Documents/MİGRATE/Python/Trade_Assistant_Files/crypto_4h_dataset.xlsx",
            "240.0_btc": "/Users/ardaerkan/Documents/MİGRATE/Python/Trade_Assistant_Files/crypto_4h_dataset.xlsx",
        }

        if name.endswith("USDT") and minute == "15.0":
            self.excel_runner(time_dict["15.0_usdt"], result)
        elif name.endswith("USDT") and minute == "15.0":
            self.excel_runner(time_dict["15.0_btc"], result)
        elif name.endswith("USDT") and minute == "60.0":
            self.excel_runner(time_dict["60.0_usdt"], result)
        elif name.endswith("USDT") and minute == "60.0":
            self.excel_runner(time_dict["60.0_btc"], result)
        elif name.endswith("USDT") and minute == "240.0":
            self.excel_runner(time_dict["240.0_usdt"], result)
        elif name.endswith("USDT") and minute == "240.0":
            self.excel_runner(time_dict["240.0_btc"], result)

        return 1

    def excel_runner(self, filename, result):
        if filename.endswith(".xlsx"):
            print("xlsx file..")
            excel_book = load_workbook(filename)
            ws = excel_book["Sheet1"]
            data = ws.values
            columns = next(data)[0:]
            df_existing = pd.DataFrame(data, columns=columns)
            last_row = len(df_existing) + 1
            print("processing xlsx..")
            for numb, (index, row) in enumerate(result.iterrows()):
                ws.cell(row=last_row + numb, column=1, value=index)
                for col_num, value in enumerate(row, 2):
                    ws.cell(row=last_row + numb, column=col_num, value=value)
            excel_book.save(filename)
            print("xlsx file success..")
        else:
            print("csv file..")
            existing_df = pd.read_csv(
                "/Users/ardaerkan/Documents/MİGRATE/Python/Trade_Assistant_Files/Bookdene.csv"
            )
            last_index = len(existing_df)
            existing_df = pd.concat([existing_df, result], axis=0)
            existing_df.to_csv(
                "/Users/ardaerkan/Documents/MİGRATE/Python/Trade_Assistant_Files/Bookdene.csv",
                mode="a",
                header=False,
            )

        return 1

    def main_executor(self, contract, start, end, coin_list):
        timeframes = ["15m", "1h", "4h"]

        results = []

        for coin in coin_list:
            subcoinlist = [coin + "USDT", coin + "BTC"]
            for subcoin in subcoinlist:
                subcoin_list = []
                with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:

                    partial_main_engine = partial(self.main, subcoin, contract)

                    futures = {
                        executor.submit(
                            partial_main_engine, tf, start=start, end=end
                        ): tf
                        for tf in timeframes
                    }

                    for future in concurrent.futures.as_completed(futures):
                        timeframe = futures[future]
                        try:
                            result = future.result()
                            if result is not None:
                                subcoin_list.append(result)
                        except Exception as exc:
                            print(
                                f"Error processing data for timeframe {timeframe}: {exc}"
                            )
                            raise
                results.append(subcoin_list)

        return self.excel_mapper(results)


if __name__ == "__main__":

    START_DATE = "2023-06-01"
    END_DATE = "2024-06-08"
    CONTRACT = "spot"
    COINS = ["AVAX"]
    data_fetcher = BinanceDataHelper()
    result = data_fetcher.main_executor(CONTRACT, START_DATE, END_DATE, COINS)
    print(result)

    # coinlist = config.get_coin_list()
